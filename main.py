from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.popup import Popup
from kivy.graphics import Color, Rectangle
from kivy.core.window import Window

import openpyxl
import os

Window.size = (360, 640)

class ExcelReaderApp(App):
    def build(self):
        self.current_row = 0
        self.current_col = 0
        self.last_row = 0
        self.last_col = 0
        self.wb = None
        self.ws = None
        
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        self.btn_select = Button(text='انتخاب فایل اکسل', size_hint=(1, 0.1))
        self.btn_select.bind(on_press=self.select_file)
        main_layout.add_widget(self.btn_select)
        
        self.color_widget = BoxLayout(size_hint=(1, 0.4))
        self.color_widget.canvas.before.add(Color(1, 1, 1, 1))
        self.color_widget.canvas.before.add(Rectangle(pos=self.color_widget.pos, size=self.color_widget.size))
        main_layout.add_widget(self.color_widget)
        
        self.lbl_value = Label(text='خالی', font_size='30sp', size_hint=(1, 0.2))
        main_layout.add_widget(self.lbl_value)
        
        self.lbl_position = Label(text='سلول: ---', font_size='16sp', size_hint=(1, 0.1))
        main_layout.add_widget(self.lbl_position)
        
        self.btn_next = Button(text='ENTER → (سلول بعدی)', size_hint=(1, 0.2))
        self.btn_next.bind(on_press=self.next_cell)
        main_layout.add_widget(self.btn_next)
        
        Window.bind(on_key_down=self.on_key_down)
        
        return main_layout
    
    def select_file(self, instance):
        filechooser = FileChooserListView()
        popup = Popup(title='انتخاب فایل اکسل', content=filechooser, size_hint=(0.9, 0.9))
        
        def load_file(chooser, selection):
            if selection:
                file_path = selection[0]
                try:
                    self.wb = openpyxl.load_workbook(file_path, data_only=True)
                    self.ws = self.wb.active
                    
                    self.last_row = self.find_last_row()
                    self.last_col = self.find_last_col()
                    
                    self.current_row = self.last_row
                    self.current_col = self.last_col
                    
                    self.show_cell()
                    popup.dismiss()
                    self.btn_select.text = f'فایل: {os.path.basename(file_path)}'
                except Exception as e:
                    self.show_error(f'خطا: {str(e)}')
        
        filechooser.bind(on_submit=load_file)
        popup.open()
    
    def find_last_row(self):
        for r in range(self.ws.max_row, 0, -1):
            for c in range(1, self.ws.max_column + 1):
                if self.ws.cell(r, c).value not in (None, "", " "):
                    return r
        return 1
    
    def find_last_col(self):
        for c in range(self.ws.max_column, 0, -1):
            for r in range(1, self.ws.max_row + 1):
                if self.ws.cell(r, c).value not in (None, "", " "):
                    return c
        return 1
    
    def get_hex_color(self, cell):
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            return "#FFFFFF"
        
        if fill.fgColor.type == "rgb" and fill.fgColor.rgb:
            rgb = fill.fgColor.rgb
            if len(rgb) == 8:
                return f"#{rgb[2:]}"
            elif len(rgb) == 6:
                return f"#{rgb}"
        
        if fill.fgColor.type == "theme":
            try:
                theme_idx = fill.fgColor.theme
                tc = self.wb.theme.themeElements.clrScheme
                arr = [tc.dk1, tc.lt1, tc.dk2, tc.lt2,
                       tc.accent1, tc.accent2, tc.accent3,
                       tc.accent4, tc.accent5, tc.accent6]
                return "#" + arr[theme_idx].srgbClr.val
            except:
                pass
        
        return "#FFFFFF"
    
    def show_cell(self):
        if self.ws is None:
            return
        
        cell = self.ws.cell(self.current_row, self.current_col)
        
        color_hex = self.get_hex_color(cell)
        r = int(color_hex[1:3], 16) / 255.0
        g = int(color_hex[3:5], 16) / 255.0
        b = int(color_hex[5:7], 16) / 255.0
        
        self.color_widget.canvas.before.clear()
        with self.color_widget.canvas.before:
            Color(r, g, b, 1)
            Rectangle(pos=self.color_widget.pos, size=self.color_widget.size)
        
        value = cell.value if cell.value is not None else "خالی"
        self.lbl_value.text = str(value)
        self.lbl_position.text = f'سلول: R{self.current_row} C{self.current_col} | رنگ: {color_hex}'
    
    def next_cell(self, instance):
        if self.ws is None:
            self.show_error('اول فایل اکسل را انتخاب کنید!')
            return
        
        self.current_col -= 1
        if self.current_col < 1:
            self.current_row -= 1
            self.current_col = self.last_col
        
        if self.current_row < 1:
            self.lbl_value.text = "پایان داده‌ها!"
            self.lbl_position.text = "به انتهای نقشه رسیدید"
            return
        
        self.show_cell()
    
    def on_key_down(self, window, key, *args):
        if key == 13:
            self.next_cell(None)
    
    def show_error(self, message):
        popup = Popup(title='خطا', content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

if __name__ == '__main__':
    ExcelReaderApp().run()
